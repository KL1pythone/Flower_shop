import logging
from amocrm.v2 import tokens
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.middlewares.logging import LoggingMiddleware
from aiogram.dispatcher import FSMContext
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
import requests
import json
import re
import tinytuya
import openpyxl
from datetime import datetime
import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import uuid
import yookassa
from yookassa import Configuration, Payment

API_TOKEN = '7264338578:AAHxUZceOZmhGCRQW8S1s9p5H-7UXTpoWGg'
YOOKASSA_SHOP_ID = '455092'
YOOKASSA_SECRET_KEY = 'live_2r-cQRDXNpI3PSe1sibeqqPq7IOMAJFGSmDD3WOuxwg'

logging.basicConfig(level=logging.INFO)

bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())
dp.middleware.setup(LoggingMiddleware())

ACCESS_ID = '57ppxafxxk5c7dcvhjku'
ACCESS_KEY = '047256cc46ea457d8d67df22a407b941'
DEVICE_ID = 'bfeba42a88ab3e74d1todd'
REGION_KEY = "eu"

device = tinytuya.Cloud(
    apiRegion=REGION_KEY,
    apiKey=ACCESS_ID,
    apiSecret=ACCESS_KEY,
    apiDeviceID=DEVICE_ID
)

new_order_button = ReplyKeyboardMarkup(
    resize_keyboard=True,
    one_time_keyboard=True
).add(KeyboardButton('Купить букет'))

class OrderFlower(StatesGroup):
    waiting_for_flower_number = State()
    waiting_for_name = State()
    waiting_for_phone_number = State()

def configure_yookassa():
    Configuration.account_id = YOOKASSA_SHOP_ID
    Configuration.secret_key = YOOKASSA_SECRET_KEY

def get_flowers_catalog():
    headers = {'Content-Type': 'application/json'}
    response = requests.get(url="https://store.tildaapi.com/api/getproductslist/?storepartuid=237506034291&recid=802627417&c=1727210376595&getparts=true&getoptions=true&slice=1&filters%5Bquantity%5D=y&sort%5Bcreated%5D=desc&size=100", headers=headers)
    if response.status_code == 200:
        data = response.json()
        products = data.get('products', [])
        flowers = {}
        for product in products:
            title = product.get('title')
            price = product.get('price')
            if title and price:
                number = re.search(r'\d+', title)
                if number:
                    title_number = number.group(0)
                    flowers[title_number] = price
        return flowers
    return {}

def open_fridge_door():
    result = device.sendcommand(DEVICE_ID, {'commands': [{'code': 'switch_1', 'value': False}]})
    time.sleep(10)
    result = device.sendcommand(DEVICE_ID, {'commands': [{'code': 'switch_1', 'value': True}]})
    return result['success']

def init_excel():
    try:
        load_workbook('orders.xlsx')
    except (FileNotFoundError, InvalidFileException):
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"
        ws.append(['ID', 'Flower Number', 'Phone Number', 'Name', 'Date Time', 'Payment Confirmed'])
        wb.save('orders.xlsx')

def save_order(flower_number, phone_number, name, payment_confirmed):
    wb = openpyxl.load_workbook('orders.xlsx')
    ws = wb.active
    orders_count = ws.max_row
    ws.append([
        orders_count,
        flower_number,
        phone_number,
        name,
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        payment_confirmed
    ])
    wb.save('orders.xlsx')

def update_payment_confirmation(flower_number):
    wb = openpyxl.load_workbook('orders.xlsx')
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == flower_number and ws.cell(row=row, column=6).value == 0:
            ws.cell(row=row, column=6).value = 1
    wb.save('orders.xlsx')

@dp.message_handler(lambda message: message.text == "Купить букет", state='*')
async def new_bouquet(message: types.Message):
    await cmd_start(message)

@dp.message_handler(commands='start', state='*')
async def cmd_start(message: types.Message):
    await message.answer("Добрый день)) введите <b>номер букета</b>", parse_mode='HTML')
    await OrderFlower.waiting_for_flower_number.set()

@dp.message_handler(state=OrderFlower.waiting_for_flower_number)
async def process_flower_number(message: types.Message, state: FSMContext):
    flowers = get_flowers_catalog()
    flower_number = message.text
    if flower_number in flowers:
        await state.update_data(flower_number=flower_number)
        await message.answer("укажите <b>ваше имя</b>", reply_markup=new_order_button, parse_mode='HTML')
        await OrderFlower.waiting_for_name.set()
    else:
        await message.answer(
            'К сожалению, этот букет уже приобрели или вы указали неверный номер, выберите пожалуйста другой номер или посмотрите те, что доступны к выбору на <a href="https://utteri.cvetoot.ru">сайте</a>',
            reply_markup=new_order_button,
            parse_mode='HTML'
        )

@dp.message_handler(state=OrderFlower.waiting_for_name)
async def process_name(message: types.Message, state: FSMContext):
    await state.update_data(name=message.text)
    await message.answer(
        'укажите свой <b>номер телефона</b> в формате +79... \n\n *отправляя вы соглашаетесь с <a href="http://cvetoot.ru/docs">офертой</a>',
        reply_markup=new_order_button,
        parse_mode='HTML'
    )
    await OrderFlower.waiting_for_phone_number.set()

async def create_payment(amount, chat_id, user_data, message):
    flower_number = user_data['flower_number']
    name = user_data['name']
    phone_number = message.text
    flowers = get_flowers_catalog()
    amount = int(float(flowers[flower_number]))

    id_key = str(uuid.uuid4())
    payment = Payment.create({
        "amount": {
            "value": f"{amount}",
            "currency": "RUB"
        },
        "payment_method_data": {
            "type": "bank_card"
        },
        "confirmation": {
            "type": "redirect",
            "return_url": "https://t.me/cvetoot_bot"
        },
        "capture": True,
        "description": f"Оплата за букет номер {flower_number}",
        "receipt": {
            "customer": {
                "full_name": name,
                "phone": phone_number
            },
            "items": [
                {
                    "description": f"Букет номер {flower_number}",
                    "quantity": "1.00",
                    "amount": {
                        "value": f"{amount}",
                        "currency": "RUB"
                    },
                    "vat_code": "1",
                    "payment_mode": "full_payment",
                    "payment_subject": "commodity"
                }
            ]
        }
    }, id_key)

    return payment

@dp.message_handler(state=OrderFlower.waiting_for_phone_number)
async def process_phone_number(message: types.Message, state: FSMContext):
    phone_number = message.text

    # Проверка номера телефона
    if not (phone_number.startswith("+79") and len(phone_number) == 12 and phone_number[1:].isdigit()):
        await message.answer("Номер телефона должен быть в формате +79 и содержать 11 цифр. Пожалуйста, введите номер заново.")
        return  # Возвращаем без завершения состояния, чтобы пользователь мог ввести номер ещё раз

    user_data = await state.get_data()
    flower_number = user_data['flower_number']
    name = user_data['name']

    payment = await create_payment(0, message.chat.id, user_data, message)

    save_order(flower_number, phone_number, name, 0)

    markup = InlineKeyboardMarkup().add(
        InlineKeyboardButton(
            text='Я оплатил букет',
            callback_data=f'check_{payment.id}'
        )
    )

    await bot.send_message(
        chat_id=message.chat.id,
        text=f"Для оплаты перейдите по <a href='{payment.confirmation.confirmation_url}'>ссылке</a>",
        parse_mode='HTML',
        reply_markup=markup
    )
    await state.finish()

@dp.callback_query_handler(lambda call: call.data.startswith('check_'))
async def check_payment(call: types.CallbackQuery, name, phone_number, amount):
    payment_id = call.data.split('_')[1]
    payment = yookassa.Payment.find_one(payment_id)

    if payment.status == 'succeeded':
        metadata = payment.metadata
        flower_number = metadata.get('flower_number', None)
        if flower_number:
            update_payment_confirmation(flower_number)
        await bot.send_message(call.message.chat.id, "Спасибо за покупку) <b>дверь открыта</b>, можете забрать букет", parse_mode='HTML')
        open_fridge_door()

        # Код из "код_1" начинается здесь
        subdomain = "cvetoot"

        tokens.default_token_manager(
            client_id="7acbc805-d521-48fa-883a-da939b17ff7c",
            client_secret="sMf6L4gzaCjeS6zGvQATDIzMcHkvNTfB6LTb2Hkgy2KCbYJyy0et9OZFPohUJ3MS",
            subdomain="cvetoot",
            redirect_url="https://t.me/cvetoot_izbata7_bot",
            storage=tokens.FileTokensStorage()
        )

        access_token = open('access_token.txt', 'r')
        refresh_token = open('refresh_token.txt', 'r')

        headers = {
            "Authorization": f"Bearer {access_token.read()}",
            "Content-Type": "application/json"
        }

        deal_data = [
            {
                "name": "Заказ из Юттери",
                "price": amount,
                "custom_fields_values": [
                    {
                        "field_id": 1387641,
                        "values": [
                            {
                                "value": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                        ]
                    },
                    {
                        "field_id": 1386075,
                        "values": [
                            {
                                "value": "Самоосблуживание (Юттери)"
                            }
                        ]
                    },
                    {
                        "field_id": 1382297,
                        "values": [
                            {
                                "value": flower_number
                            }
                        ]
                    }
                ]
            }
        ]

        deal_response = requests.post(f"https://{subdomain}.amocrm.ru/api/v4/leads", headers=headers, data=json.dumps(deal_data))
        deal = deal_response.json().get("_embedded", {}).get("leads", [{}])[0]
        deal_id = deal.get("id")

        contact_data = [
            {
                "name": name,
                "custom_fields_values": [
                    {
                        "field_code": "PHONE",
                        "values": [
                            {
                                "value": phone_number,
                                "enum_code": "WORK"
                            }
                        ]
                    }
                ]
            }
        ]

        contact_response = requests.post(f"https://{subdomain}.amocrm.ru/api/v4/contacts", headers=headers, data=json.dumps(contact_data))
        contact = contact_response.json().get("_embedded", {}).get("contacts", [{}])[0]
        contact_id = contact.get("id")

        link_data = [
            {
                "to_entity_id": contact_id,
                "to_entity_type": "contacts",
                "metadata": {
                    "is_main": True
                }
            }
        ]

        link_response = requests.post(f"https://{subdomain}.amocrm.ru/api/v4/leads/{deal_id}/link", headers=headers, data=json.dumps(link_data))

    else:
        await bot.send_message(call.message.chat.id, "Оплата не завершена. Пожалуйста, проверьте статус платежа.")

if __name__ == '__main__':
    configure_yookassa()
    init_excel()
    from aiogram import executor
    executor.start_polling(dp, skip_updates=True)